from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import random
import shutil
import subprocess
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any


VIDEO_SUFFIXES = {".mp4", ".mov", ".mkv", ".avi"}
CACHE_VERSION = 1
MANIFEST_COLUMNS = [
    "mode",
    "output_name",
    "source_video",
    "audio_source",
    "variant_id",
    "deleted_frames",
    "cover_timestamp",
    "cover_quality_status",
    "cover_similarity_status",
    "combo_signature",
    "source_chain",
    "variation_status",
    "duplicate_risk",
    "quality_status",
    "business_tag",
    "material_type",
    "authorization_note",
    "upload_note",
    "creative_unit_id",
    "note",
]


@dataclass
class SourceVideo:
    path: Path
    duration: float
    width: int
    height: int
    fps: float
    frame_count: int
    has_audio: bool
    pix_fmt: str = ""
    bit_rate: int = 0
    md5: str = ""


@dataclass
class CoverCandidate:
    timestamp: float
    path: Path
    sharpness: float = 0.0
    brightness: float = 0.0
    hash_value: str = ""


@dataclass
class VariantPlan:
    variant_id: str
    source_video: Path
    output_name: str
    output_path: Path
    mode: str = "frame_variation"
    deleted_frames: list[int] = field(default_factory=list)
    frame_keep_expr: str = "1"
    frame_signature: str = ""
    combo_signature: str = ""
    source_chain: list[Path] = field(default_factory=list)
    cover_timestamp: float = 0.0
    cover_quality_status: str = "pending"
    cover_similarity_status: str = "pending"
    quality_status: str = "ready"
    variation_status: str = "unique"
    duplicate_risk: str = "low"
    business_tag: str = ""
    material_type: str = ""
    authorization_note: str = ""
    upload_note: str = ""
    creative_unit_id: str = ""
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["source_video"] = str(self.source_video)
        data["output_path"] = str(self.output_path)
        data["source_chain"] = [str(path) for path in self.source_chain]
        return data


def _file_cache_key(path: Path) -> str:
    stat = path.stat()
    return f"{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}"


def _load_cache(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, dict) or payload.get("version") != CACHE_VERSION:
        return {}
    entries = payload.get("entries")
    return dict(entries) if isinstance(entries, dict) else {}


def _save_cache(path: Path, entries: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f"{path.name}.{os.getpid()}.{threading.get_ident()}.part")
    temporary.write_text(
        json.dumps({"version": CACHE_VERSION, "entries": entries}, ensure_ascii=False),
        encoding="utf-8",
    )
    os.replace(temporary, path)


def _source_cache_value(source: SourceVideo) -> dict[str, Any]:
    return {
        "duration": source.duration,
        "width": source.width,
        "height": source.height,
        "fps": source.fps,
        "frame_count": source.frame_count,
        "has_audio": source.has_audio,
        "pix_fmt": source.pix_fmt,
        "bit_rate": source.bit_rate,
        "md5": source.md5,
    }


def _source_from_cache(path: Path, value: dict[str, Any]) -> SourceVideo:
    return SourceVideo(
        path=path,
        duration=float(value.get("duration", 0.0) or 0.0),
        width=int(value.get("width", 0) or 0),
        height=int(value.get("height", 0) or 0),
        fps=float(value.get("fps", 0.0) or 0.0),
        frame_count=int(value.get("frame_count", 0) or 0),
        has_audio=bool(value.get("has_audio")),
        pix_fmt=str(value.get("pix_fmt", "") or ""),
        bit_rate=int(value.get("bit_rate", 0) or 0),
        md5=str(value.get("md5", "") or ""),
    )


def _cover_cache_key(video_path: Path, duration: float, variant_index: int, attempt: int) -> str:
    return "|".join(
        (
            _file_cache_key(video_path),
            f"{duration:.6f}",
            str(variant_index),
            str(attempt),
        )
    )


class _NoopLock:
    def __enter__(self) -> "_NoopLock":
        return self

    def __exit__(self, *_: object) -> None:
        return None


def main() -> int:
    args = parse_args()
    task_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_root = Path(args.output_root).resolve()
    task_root = output_root / task_id
    videos_dir = task_root / "videos"
    temp_cover_dir = task_root / "_cover_candidates"
    videos_dir.mkdir(parents=True, exist_ok=True)
    temp_cover_dir.mkdir(parents=True, exist_ok=True)
    run_log = task_root / "run.log"
    log_event(run_log, f"start frame_variation task_id={task_id}")
    log_event(run_log, f"config={json.dumps(vars(args), ensure_ascii=False)}")

    variants: list[VariantPlan] = []

    try:
        ffmpeg = resolve_binary("ffmpeg", args.ffmpeg)
        ffprobe = resolve_binary("ffprobe", args.ffprobe)
        vfr_args = detect_vfr_args(ffmpeg)
        source_paths = [Path(item).resolve() for item in args.source]
        for path in source_paths:
            if not path.is_file():
                raise SystemExit(f"Source video does not exist: {path}")
            if path.suffix.lower() not in VIDEO_SUFFIXES:
                raise SystemExit(f"Unsupported video suffix: {path}")

        selected_covers: list[CoverCandidate] = []
        used_names: set[str] = set()
        total = len(source_paths) * args.target_count
        cache_root = output_root.parent
        source_cache_path = cache_root / ".source_metadata_cache.json"
        cover_cache_path = cache_root / ".cover_candidate_cache.json"
        source_cache = _load_cache(source_cache_path)
        cover_cache = _load_cache(cover_cache_path)
        source_cache_lock = threading.Lock()
        cover_cache_lock = threading.Lock()
        selected_covers_lock = threading.Lock()
        used_names_lock = threading.Lock()
        log_lock = threading.Lock()
        base_seed = args.seed or task_id

        def log_threadsafe(message: str) -> None:
            with log_lock:
                log_event(run_log, message)

        def process_source(source_index: int, source_path: Path) -> tuple[int, list[VariantPlan], BaseException | None]:
            completed: list[VariantPlan] = []
            try:
                log_threadsafe(f"source_start index={source_index} path={source_path}")
                source = load_source_video(
                    source_path,
                    ffprobe,
                    cache=source_cache,
                    cache_path=source_cache_path,
                    cache_lock=source_cache_lock,
                )
                used_signatures: set[str] = set()
                rng = random.Random(f"{base_seed}:{source_index}")
                max_retries = max(args.max_retries, args.target_count * 4)
                for item_index in range(args.target_count):
                    completed.append(
                        build_one_variant(
                            source=source,
                            source_index=source_index,
                            item_index=item_index,
                            total=total,
                            videos_dir=videos_dir,
                            temp_cover_dir=temp_cover_dir,
                            selected_covers=selected_covers,
                            selected_covers_lock=selected_covers_lock,
                            used_signatures=used_signatures,
                            used_names=used_names,
                            used_names_lock=used_names_lock,
                            rng=rng,
                            max_retries=max_retries,
                            args=args,
                            ffmpeg=ffmpeg,
                            cover_cache=cover_cache,
                            cover_cache_path=cover_cache_path,
                            cover_cache_lock=cover_cache_lock,
                            render=not args.skip_cover or args.plan_only,
                        )
                    )
                if args.skip_cover and not args.plan_only:
                    for batch_start in range(0, len(completed), max(1, args.variant_batch_size)):
                        batch = completed[batch_start:batch_start + max(1, args.variant_batch_size)]
                        render_frame_drop_variants_batched(
                            source.path,
                            batch,
                            args.width,
                            args.height,
                            args.resize_mode,
                            ffmpeg,
                            has_audio=source.has_audio,
                            source_width=source.width,
                            source_height=source.height,
                            source_fps=source.fps,
                            source_pix_fmt=source.pix_fmt,
                            vfr_args=vfr_args,
                        )
                        for batch_index, variant in enumerate(batch, start=batch_start + 1):
                            print(f"[{batch_index}/{args.target_count}] {variant.output_path}", flush=True)
                return source_index, completed, None
            except BaseException as exc:
                return source_index, completed, exc

        completed_by_source: dict[int, list[VariantPlan]] = {}
        errors: list[BaseException] = []
        max_workers = min(max(1, int(args.source_concurrency)), len(source_paths))
        with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="frame-fission") as executor:
            futures = [
                executor.submit(process_source, source_index, source_path)
                for source_index, source_path in enumerate(source_paths, start=1)
            ]
            completed_count = 0
            for future in as_completed(futures):
                source_index, source_variants, error = future.result()
                completed_by_source[source_index] = source_variants
                for variant in source_variants:
                    completed_count += 1
                    log_threadsafe(
                        f"variant_done {completed_count}/{total} id={variant.variant_id} output={variant.output_path}"
                    )
                    if not args.skip_cover:
                        print(f"[{completed_count}/{total}] {variant.output_path}", flush=True)
                if error is not None:
                    errors.append(error)

        variants = [
            variant
            for source_index in sorted(completed_by_source)
            for variant in completed_by_source[source_index]
        ]
        if errors:
            raise errors[0]
        payload = {
            "task_id": task_id,
            "mode": "frame_variation_plan" if args.plan_only else "frame_variation",
            "status": "success",
            "rendered": not args.plan_only,
            "config": vars(args),
            "source_videos": [str(path) for path in source_paths],
            "variants": [item.to_dict() for item in variants],
        }
        write_outputs(task_root, payload, variants)
        log_event(run_log, f"success completed={len(variants)} task_root={task_root}")
        print(str(task_root))
        return 0
    except BaseException as exc:
        write_error(task_root, "frame_variation", args, variants, exc)
        raise
    finally:
        shutil.rmtree(temp_cover_dir, ignore_errors=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate frame-drop video variants.")
    parser.add_argument("--source", nargs="+", required=True, help="One or more source videos.")
    parser.add_argument("--output-root", required=True, help="Directory that will receive a timestamped task folder.")
    parser.add_argument("--task-name", default="campaign_batch")
    parser.add_argument("--target-count", type=int, default=20, help="Variants per source video.")
    parser.add_argument(
        "--source-concurrency",
        "--concurrency",
        dest="source_concurrency",
        type=int,
        default=3,
        help="Maximum number of source videos rendered concurrently.",
    )
    parser.add_argument(
        "--variant-batch-size",
        type=int,
        default=4,
        help="Number of no-cover variants rendered by one FFmpeg process.",
    )
    parser.add_argument("--frames-per-second-drop", type=int, default=1)
    parser.add_argument("--width", type=int, default=720)
    parser.add_argument("--height", type=int, default=1280)
    parser.add_argument("--resize-mode", choices=("crop", "contain", "stretch", "original"), default="crop")
    parser.add_argument("--cover-hold-seconds", type=float, default=0.8)
    parser.add_argument(
        "--skip-cover",
        action="store_true",
        help="只生成抽帧变体，不抽取或拼接封面前贴。",
    )
    parser.add_argument("--max-retries", type=int, default=120)
    parser.add_argument("--max-name-length", type=int, default=240)
    parser.add_argument("--seed", default="")
    parser.add_argument("--business-tag", default="")
    parser.add_argument("--material-type", default="")
    parser.add_argument("--authorization-note", default="")
    parser.add_argument("--upload-note", default="")
    parser.add_argument("--ffmpeg", default="")
    parser.add_argument("--ffprobe", default="")
    parser.add_argument(
        "--plan-only",
        action="store_true",
        help="只生成删帧和封面计划，不渲染中间底片视频。",
    )
    return parser.parse_args()


def resolve_binary(name: str, explicit: str = "") -> str:
    candidates = [
        explicit,
        os.environ.get(f"{name.upper()}_BIN", ""),
        shutil.which(name) or "",
        shutil.which(f"{name}.exe") or "",
        str(Path(__file__).resolve().parent / "bin" / f"{name}.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(Path(candidate))
    raise SystemExit(f"Missing dependency: {name}. Pass --{name} or set {name.upper()}_BIN.")


def detect_vfr_args(ffmpeg: str) -> list[str]:
    """选择当前 FFmpeg 支持的可变帧率参数，兼容新旧版本。"""
    try:
        result = subprocess.run(
            [ffmpeg, "-hide_banner", "-h", "full"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=10,
        )
    except (OSError, subprocess.SubprocessError):
        return []
    help_text = f"{result.stdout}\n{result.stderr}"
    if "fps_mode" in help_text:
        return ["-fps_mode", "vfr"]
    if "vsync" in help_text:
        return ["-vsync", "0"]
    return []


def run_process(args: list[str]) -> subprocess.CompletedProcess[str]:
    startupinfo = None
    creationflags = 0
    if hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        creationflags = subprocess.CREATE_NO_WINDOW
    try:
        return subprocess.run(
            args,
            check=True,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            encoding="utf-8",
            errors="replace",
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except subprocess.CalledProcessError as exc:
        raise RuntimeError((exc.stderr or exc.stdout or str(exc)).strip()) from exc


def load_source_video(
    path: Path,
    ffprobe: str,
    *,
    cache: dict[str, Any] | None = None,
    cache_path: Path | None = None,
    cache_lock: threading.Lock | None = None,
) -> SourceVideo:
    cache_key = _file_cache_key(path)
    if cache is not None:
        lock = cache_lock or _NoopLock()
        with lock:
            cached = cache.get(cache_key)
        if isinstance(cached, dict) and cached.get("pix_fmt"):
            return _source_from_cache(path, cached)

    result = run_process([ffprobe, "-v", "error", "-print_format", "json", "-show_streams", "-show_format", str(path)])
    data = json.loads(result.stdout)
    fmt = data.get("format", {})
    video = next((item for item in data.get("streams", []) if item.get("codec_type") == "video"), {})
    audio = next((item for item in data.get("streams", []) if item.get("codec_type") == "audio"), {})
    fps = fraction_to_float(video.get("avg_frame_rate", "0/1"))
    duration = float(fmt.get("duration", 0.0) or 0.0)
    frames = int(video.get("nb_frames", 0) or round(duration * fps))
    source = SourceVideo(
        path=path,
        duration=duration,
        width=int(video.get("width", 0) or 0),
        height=int(video.get("height", 0) or 0),
        fps=fps,
        frame_count=frames,
        has_audio=bool(audio),
        pix_fmt=str(video.get("pix_fmt", "") or ""),
        bit_rate=int(fmt.get("bit_rate", 0) or 0),
        md5=compute_md5(path),
    )
    if cache is not None:
        lock = cache_lock or _NoopLock()
        with lock:
            cache[cache_key] = _source_cache_value(source)
            if cache_path is not None:
                _save_cache(cache_path, cache)
    return source


def fraction_to_float(value: str) -> float:
    if "/" not in value:
        return float(value or 0.0)
    left, right = value.split("/", 1)
    denominator = float(right or 1.0)
    return float(left or 0.0) / denominator if denominator else 0.0


def compute_md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_one_variant(
    *,
    source: SourceVideo,
    source_index: int,
    item_index: int,
    total: int,
    videos_dir: Path,
    temp_cover_dir: Path,
    selected_covers: list[CoverCandidate],
    used_signatures: set[str],
    used_names: set[str],
    rng: random.Random,
    max_retries: int,
    args: argparse.Namespace,
    ffmpeg: str,
    selected_covers_lock: threading.Lock | None = None,
    used_names_lock: threading.Lock | None = None,
    cover_cache: dict[str, Any] | None = None,
    cover_cache_path: Path | None = None,
    cover_cache_lock: threading.Lock | None = None,
    render: bool = True,
) -> VariantPlan:
    for attempt in range(max_retries):
        deleted_frames, frame_keep_expr = choose_compact_frame_drop_plan(
            source.fps,
            source.frame_count,
            args.frames_per_second_drop,
            rng,
        )
        signature = ",".join(str(frame) for frame in sorted(deleted_frames))
        if signature in used_signatures:
            continue

        cover: CoverCandidate | None = None
        quality = "skipped"
        similarity = "skipped"
        cover_lock = selected_covers_lock or _NoopLock()
        if not args.skip_cover:
            cover_name = (
                f"{sanitize_windows_stem(source.path.stem, args.max_name_length)}"
                f"_cover_{source_index}_{item_index + 1}_{attempt}"
            )
            with cover_lock:
                cover, quality, similarity = choose_cover(
                    source.path,
                    temp_cover_dir,
                    cover_name,
                    selected_covers,
                    source.duration,
                    item_index,
                    attempt,
                    ffmpeg,
                    cache=cover_cache,
                    cache_path=cover_cache_path,
                    cache_lock=cover_cache_lock,
                )
            if similarity != "unique" and attempt < 3:
                continue
            if similarity != "unique":
                similarity = f"{similarity}_fallback"

        name_lock = used_names_lock or _NoopLock()
        with name_lock:
            output_name = unique_output_name(source.path.stem, item_index + 1, args.max_name_length, used_names)
        output_path = videos_dir / f"{output_name}.mp4"
        if cover is not None:
            with cover_lock:
                selected_covers.append(cover)

        try:
            if not args.plan_only and render:
                render_frame_drop_variant_with_cover(
                    source.path,
                    output_path,
                    deleted_frames,
                    cover.timestamp if cover is not None else 0.0,
                    args.cover_hold_seconds,
                    args.width,
                    args.height,
                    args.resize_mode,
                    ffmpeg,
                    has_audio=source.has_audio,
                    frame_keep_expr=frame_keep_expr,
                    cover_path=cover.path if cover is not None else None,
                    include_cover=not args.skip_cover,
                    source_width=source.width,
                    source_height=source.height,
                    source_fps=source.fps,
                    source_pix_fmt=source.pix_fmt,
                )
        except BaseException:
            with cover_lock:
                if cover is not None and cover in selected_covers:
                    selected_covers.remove(cover)
            raise
        used_signatures.add(signature)
        global_index = (source_index - 1) * args.target_count + item_index + 1
        return VariantPlan(
            variant_id=f"VAR-{global_index:03d}",
            source_video=source.path,
            output_name=output_name,
            output_path=output_path,
            deleted_frames=deleted_frames,
            frame_keep_expr=frame_keep_expr,
            frame_signature=signature,
            combo_signature=signature,
            cover_timestamp=cover.timestamp if cover is not None else 0.0,
            cover_quality_status=quality,
            cover_similarity_status=similarity,
            business_tag=args.business_tag,
            material_type=args.material_type,
            authorization_note=args.authorization_note,
            upload_note=args.upload_note,
        )
    raise RuntimeError(f"Could not find a unique frame-drop plan for {source.path.name} variant {item_index + 1}/{total}.")


def choose_deleted_frames_per_second(fps: float, frame_count: int, frames_per_second: int, rng: random.Random) -> list[int]:
    if fps <= 0 or frame_count <= 0:
        return []
    deleted: list[int] = []
    seconds = max(1, int(frame_count / fps))
    picks = max(1, frames_per_second)
    for second in range(seconds):
        start = int(round(second * fps))
        end = min(frame_count, int(round((second + 1) * fps)))
        # 保留首尾帧，避免删帧改变主视频的时间原点或结尾时间。
        start = max(start, 1)
        end = min(end, frame_count - 1)
        if end <= start:
            continue
        candidates = list(range(start, end))
        deleted.extend(rng.sample(candidates, k=min(picks, len(candidates))))
    return sorted(set(deleted))


def choose_compact_frame_drop_plan(
    fps: float,
    frame_count: int,
    frames_per_second: int,
    rng: random.Random,
) -> tuple[list[int], str]:
    """Plan per-second frame drops with a constant-size FFmpeg expression."""
    if fps <= 0 or frame_count <= 2:
        return [], "1"
    frame_group = max(2, int(round(fps)))
    drop_count = min(max(1, frames_per_second), frame_group - 1)
    multipliers = [
        value
        for value in range(1, frame_group)
        if math.gcd(value, frame_group) == 1
    ]
    multiplier = rng.choice(multipliers or [1])
    seed = rng.randrange(1, frame_group)
    deleted_frames: list[int] = []
    for frame in range(1, frame_count - 1):
        second = frame // frame_group
        selected_offset = (second * multiplier + seed) % frame_group
        if (frame % frame_group - selected_offset) % frame_group < drop_count:
            deleted_frames.append(frame)

    offset_expr = (
        f"mod(mod(n\\,{frame_group})-mod(floor(n/{frame_group})*{multiplier}+{seed}\\,{frame_group})"
        f"+{frame_group}\\,{frame_group})"
    )
    keep_expr = f"1-lt({offset_expr}\\,{drop_count})*gt(n\\,0)*lt(n\\,{frame_count - 1})"
    return deleted_frames, keep_expr


def unique_output_name(stem: str, sequence: int, max_length: int, used_names: set[str]) -> str:
    suffix = f"({sequence})"
    base = sanitize_windows_stem(stem, max(1, max_length - len(suffix)))
    candidate = f"{base}{suffix}"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    offset = 2
    while True:
        alt_suffix = f"({sequence})-{offset:02d}"
        candidate = f"{sanitize_windows_stem(stem, max(1, max_length - len(alt_suffix)))}{alt_suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        offset += 1


def sanitize_windows_stem(name: str, max_length: int) -> str:
    invalid = set('<>:"/\\|?*')
    cleaned = "".join("_" if ch in invalid or ord(ch) < 32 else ch for ch in name.strip())
    cleaned = " ".join(cleaned.split())
    return cleaned[:max_length].rstrip(" ._-") or "variant"


def choose_cover(
    video_path: Path,
    temp_dir: Path,
    output_name: str,
    existing: list[CoverCandidate],
    duration: float,
    variant_index: int,
    attempt: int,
    ffmpeg: str,
    *,
    cache: dict[str, Any] | None = None,
    cache_path: Path | None = None,
    cache_lock: threading.Lock | None = None,
) -> tuple[CoverCandidate, str, str]:
    usable = max(0.5, duration - 0.4)
    base = ((variant_index * 0.61803398875 + attempt * 0.17320508075) % 1.0) * usable
    timestamps = [min(max(0.2, base + offset), usable) for offset in (0.0, 0.7, 1.4)]
    cache_key = _cover_cache_key(video_path, duration, variant_index, attempt)
    candidates: list[CoverCandidate] = []
    cache_guard = cache_lock or _NoopLock()
    if cache is not None:
        with cache_guard:
            cached = cache.get(cache_key)
        if isinstance(cached, list):
            cache_complete = True
            for index, item in enumerate(cached):
                if not isinstance(item, dict):
                    continue
                candidate_path = temp_dir / f"{output_name}_{index}.jpg"
                if not candidate_path.is_file():
                    cache_complete = False
                    break
                candidates.append(
                    CoverCandidate(
                        timestamp=float(item.get("timestamp", 0.0) or 0.0),
                        path=candidate_path,
                        sharpness=float(item.get("sharpness", 0.0) or 0.0),
                        brightness=float(item.get("brightness", 0.0) or 0.0),
                        hash_value=str(item.get("hash_value", "") or ""),
                    )
                )
            if not cache_complete:
                candidates.clear()

    if not candidates:
        for index, timestamp in enumerate(dict.fromkeys(round(value, 3) for value in timestamps)):
            path = temp_dir / f"{output_name}_{index}.jpg"
            extract_frame(video_path, timestamp, path, ffmpeg)
            candidates.append(evaluate_cover(path, timestamp))
        if cache is not None:
            cache_value = [
                {
                    "timestamp": item.timestamp,
                    "sharpness": item.sharpness,
                    "brightness": item.brightness,
                    "hash_value": item.hash_value,
                }
                for item in candidates
            ]
            with cache_guard:
                cache[cache_key] = cache_value
                if cache_path is not None:
                    _save_cache(cache_path, cache)

    ranked = sorted(candidates, key=lambda item: (item.sharpness, -abs(128 - item.brightness)), reverse=True)
    selected = ranked[0]
    selected_quality, selected_similarity = cover_status(selected, existing)
    for candidate in ranked:
        quality, similarity = cover_status(candidate, existing)
        if quality in {"clear", "not_evaluated"} and similarity == "unique":
            selected = candidate
            selected_quality = quality
            selected_similarity = similarity
            break
    for candidate in candidates:
        if candidate.path != selected.path:
            candidate.path.unlink(missing_ok=True)
    return selected, selected_quality, selected_similarity


def extract_frame(video_path: Path, timestamp: float, output_path: Path, ffmpeg: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_process([ffmpeg, "-y", "-ss", f"{timestamp:.3f}", "-i", str(video_path), "-frames:v", "1", "-vf", "scale='min(720,iw)':-2", str(output_path)])


def evaluate_cover(path: Path, timestamp: float) -> CoverCandidate:
    try:
        from PIL import Image, ImageStat
    except ImportError:
        return CoverCandidate(timestamp=timestamp, path=path, hash_value="")

    image = Image.open(path)
    grayscale = image.convert("L")
    pixels = list(grayscale.tobytes())
    width, height = grayscale.size
    diffs: list[int] = []
    for y in range(height):
        row = y * width
        for x in range(width - 1):
            diffs.append(pixels[row + x + 1] - pixels[row + x])
    for y in range(height - 1):
        row = y * width
        next_row = (y + 1) * width
        for x in range(width):
            diffs.append(pixels[next_row + x] - pixels[row + x])
    mean = sum(diffs) / len(diffs) if diffs else 0.0
    sharpness = sum((value - mean) ** 2 for value in diffs) / len(diffs) if diffs else 0.0
    brightness = float(ImageStat.Stat(grayscale).mean[0])
    return CoverCandidate(timestamp=timestamp, path=path, sharpness=sharpness, brightness=brightness, hash_value=average_hash(image))


def average_hash(image: Any, hash_size: int = 8) -> str:
    grayscale = image.convert("L").resize((hash_size, hash_size))
    pixels = list(grayscale.tobytes())
    threshold = sum(pixels) / len(pixels)
    bits = "".join("1" if pixel > threshold else "0" for pixel in pixels)
    return f"{int(bits, 2):016x}"


def cover_status(candidate: CoverCandidate, existing: list[CoverCandidate]) -> tuple[str, str]:
    if not candidate.hash_value:
        return "not_evaluated", "unique"
    quality = "clear"
    if candidate.sharpness < 12:
        quality = "blur_risk"
    if candidate.brightness < 20 or candidate.brightness > 245:
        quality = "exposure_risk"
    similarity = "unique"
    for item in existing:
        if item.hash_value and hash_distance(candidate.hash_value, item.hash_value) < 6:
            similarity = "too_similar"
            break
    return quality, similarity


def hash_distance(left: str, right: str) -> int:
    return (int(left, 16) ^ int(right, 16)).bit_count()


def output_video_filter(
    width: int,
    height: int,
    mode: str,
    *,
    source_width: int = 0,
    source_height: int = 0,
    source_fps: float = 0.0,
    source_pix_fmt: str = "",
) -> str:
    width = max(2, int(width or 0))
    height = max(2, int(height or 0))
    geometry_matches = (
        source_width > 0
        and source_height > 0
        and source_width == width
        and source_height == height
    )
    # 基础抽帧保留源时间轴。把所有素材强制转成 30 CFR 会额外生成一轮完整帧，
    # 与稀疏删帧的目标相悖；包装分支再按成品要求做时间轴归一化。
    tail_parts = ["setsar=1"]
    if str(source_pix_fmt or "").lower() != "yuv420p":
        tail_parts.append("format=yuv420p")
    tail = ",".join(tail_parts)
    if mode == "original" or geometry_matches:
        return tail
    if mode == "stretch":
        return f"scale={width}:{height},{tail}"
    if mode == "contain":
        return f"scale={width}:{height}:force_original_aspect_ratio=decrease,pad={width}:{height}:(ow-iw)/2:(oh-ih)/2:black,{tail}"
    return f"scale={width}:{height}:force_original_aspect_ratio=increase,crop={width}:{height},{tail}"


def render_frame_drop_variant_with_cover(
    source_path: Path,
    output_path: Path,
    deleted_frames: list[int],
    cover_timestamp: float,
    hold_seconds: float,
    width: int,
    height: int,
    resize_mode: str,
    ffmpeg: str,
    has_audio: bool = True,
    frame_keep_expr: str | None = None,
    *,
    cover_path: Path | None = None,
    include_cover: bool = True,
    source_width: int = 0,
    source_height: int = 0,
    source_fps: float = 0.0,
    source_pix_fmt: str = "",
    vfr_args: list[str] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if frame_keep_expr:
        drop_expr = frame_keep_expr
    else:
        conditions = "+".join(f"eq(n\\,{frame})" for frame in sorted(set(deleted_frames)))
        drop_expr = f"not({conditions})" if conditions else "1"
    hold_seconds = max(0.001, float(hold_seconds))
    use_cover_image = bool(include_cover and cover_path and cover_path.is_file())
    cover_input_index = 1 if use_cover_image else 0
    main_input_index = 0 if use_cover_image else 1
    if not include_cover:
        filter_parts = [
            # 保留选中帧的原始 PTS，避免缩短时间轴，这样未处理的音频可以直接复用。
            f"[0:v:0]select='{drop_expr}'[main]",
            f"[main]{output_video_filter(width, height, resize_mode, source_width=source_width, source_height=source_height, source_fps=source_fps, source_pix_fmt=source_pix_fmt)}[vout]",
        ]
    elif use_cover_image:
        # The cover candidates are intentionally cached at a small preview
        # size. Normalize them to the source stream geometry before concat;
        # concat rejects inputs whose dimensions differ even when the final
        # output filter would resize both inputs later.
        cover_width = max(2, int(source_width or width))
        cover_height = max(2, int(source_height or height))
        cover_filter = (
            f"[{cover_input_index}:v:0]scale={cover_width}:{cover_height},format=yuv420p,"
            f"trim=duration={hold_seconds:.3f},setpts=PTS-STARTPTS[cover]"
        )
        filter_parts = [
            cover_filter,
            f"[{main_input_index}:v:0]select='{drop_expr}',setpts=PTS-STARTPTS[main]",
            "[cover][main]concat=n=2:v=1:a=0[merged]",
            f"[merged]{output_video_filter(width, height, resize_mode, source_width=source_width, source_height=source_height, source_fps=source_fps, source_pix_fmt=source_pix_fmt)}[vout]",
        ]
    else:
        cover_filter = (
            f"[{cover_input_index}:v:0]trim=end_frame=1,setpts=PTS-STARTPTS,"
            f"tpad=stop_mode=clone:stop_duration={hold_seconds:.3f},"
            f"trim=duration={hold_seconds:.3f}[cover]"
        )
        filter_parts = [
            cover_filter,
            f"[{main_input_index}:v:0]select='{drop_expr}',setpts=PTS-STARTPTS[main]",
            "[cover][main]concat=n=2:v=1:a=0[merged]",
            f"[merged]{output_video_filter(width, height, resize_mode, source_width=source_width, source_height=source_height, source_fps=source_fps, source_pix_fmt=source_pix_fmt)}[vout]",
        ]
    if has_audio and include_cover:
        delay_ms = max(1, int(round(hold_seconds * 1000)))
        filter_parts.append(f"[{main_input_index}:a:0]asetpts=PTS-STARTPTS,adelay={delay_ms}:all=1[aout]")
    filter_complex = ";".join(filter_parts)

    command = [ffmpeg, "-y"]
    if not include_cover:
        command.extend(["-i", str(source_path)])
    elif use_cover_image:
        command.extend([
            "-i",
            str(source_path),
            "-loop",
            "1",
            "-framerate",
            "30",
            "-i",
            str(cover_path),
        ])
    else:
        command.extend([
            "-ss",
            f"{max(0.0, cover_timestamp):.3f}",
            "-i",
            str(source_path),
            "-i",
            str(source_path),
        ])
    command.extend([
        "-filter_complex",
        filter_complex,
        "-map",
        "[vout]",
        "-c:v",
        "libx264",
        "-preset",
        "ultrafast",
        "-crf",
        "24",
    ])
    if has_audio:
        if include_cover:
            command.extend(["-map", "[aout]", "-c:a", "aac", "-b:a", "192k", "-shortest"])
        else:
            # Frame deletion does not touch the audio stream.  Preserve the
            # original AAC/codec instead of decoding and encoding it again.
            command.extend(["-map", "0:a:0", "-c:a", "copy", "-shortest"])
    command.extend([*(vfr_args if vfr_args is not None else detect_vfr_args(ffmpeg)), str(output_path)])
    run_process(command)


def render_frame_drop_variants_batched(
    source_path: Path,
    variants: list[VariantPlan],
    width: int,
    height: int,
    resize_mode: str,
    ffmpeg: str,
    *,
    has_audio: bool,
    source_width: int,
    source_height: int,
    source_fps: float,
    source_pix_fmt: str,
    vfr_args: list[str] | None = None,
) -> None:
    """从一次源视频解码中批量生成多个无封面变体。

    旧逻辑每个产物启动一个 FFmpeg，长视频会被重复解复用和解码。这里让每个
    分支保留自己的删帧方案和编码器，但共享一次源视频解码；音频对每个产物
    都直接流复制。
    """
    if not variants:
        return
    temporary_paths = [
        variant.output_path.with_name(f"{variant.output_path.stem}.part{variant.output_path.suffix}")
        for variant in variants
    ]
    for path in temporary_paths:
        path.unlink(missing_ok=True)
    filter_parts = []
    for index, variant in enumerate(variants):
        filter_parts.append(
            f"[0:v:0]select='{variant.frame_keep_expr}',"
            f"{output_video_filter(width, height, resize_mode, source_width=source_width, source_height=source_height, source_fps=source_fps, source_pix_fmt=source_pix_fmt)}"
            f"[v{index}]"
        )
    command = [
        ffmpeg,
        "-hide_banner",
        "-y",
        "-i",
        str(source_path),
        "-filter_complex",
        ";".join(filter_parts),
    ]
    for index, temporary in enumerate(temporary_paths):
        command.extend([
            "-map",
            f"[v{index}]",
            "-c:v",
            "libx264",
            "-preset",
            "ultrafast",
            "-crf",
            "24",
            *(vfr_args if vfr_args is not None else detect_vfr_args(ffmpeg)),
        ])
        if has_audio:
            command.extend(["-map", "0:a:0", "-c:a", "copy", "-shortest"])
        else:
            command.append("-an")
        command.append(str(temporary))
    try:
        run_process(command)
        for temporary, variant in zip(temporary_paths, variants):
            if not temporary.is_file() or temporary.stat().st_size <= 0:
                raise RuntimeError(f"FFmpeg 未生成裂变产物：{variant.output_path.name}")
            os.replace(temporary, variant.output_path)
    except BaseException:
        for path in temporary_paths:
            path.unlink(missing_ok=True)
        raise


def write_outputs(task_root: Path, payload: dict[str, Any], variants: list[VariantPlan]) -> None:
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "task.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_manifest_csv(task_root / "manifest.csv", variants)
    write_manifest_xlsx(task_root / "manifest.xlsx", variants)
    write_task_log(task_root / "task.log", payload)


def write_manifest_csv(path: Path, variants: list[VariantPlan]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        writer.writerows(plan_to_row(item) for item in variants)


def write_manifest_xlsx(path: Path, variants: list[VariantPlan]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "manifest"
    sheet.append(MANIFEST_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in variants:
        row = plan_to_row(item)
        sheet.append([row[column] for column in MANIFEST_COLUMNS])
    workbook.save(path)


def plan_to_row(plan: VariantPlan) -> dict[str, str]:
    return {
        "mode": plan.mode,
        "output_name": plan.output_name,
        "source_video": str(plan.source_video),
        "audio_source": "",
        "variant_id": plan.variant_id,
        "deleted_frames": ",".join(str(frame) for frame in plan.deleted_frames),
        "cover_timestamp": f"{plan.cover_timestamp:.3f}",
        "cover_quality_status": plan.cover_quality_status,
        "cover_similarity_status": plan.cover_similarity_status,
        "combo_signature": plan.combo_signature,
        "source_chain": " + ".join(str(path) for path in plan.source_chain),
        "variation_status": plan.variation_status,
        "duplicate_risk": plan.duplicate_risk,
        "quality_status": plan.quality_status,
        "business_tag": plan.business_tag,
        "material_type": plan.material_type,
        "authorization_note": plan.authorization_note,
        "upload_note": plan.upload_note,
        "creative_unit_id": plan.creative_unit_id,
        "note": plan.note,
    }


def write_task_log(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        f"task_id: {payload.get('task_id', '')}",
        "mode: frame_variation",
        f"status: {payload.get('status', '')}",
        f"created_at: {datetime.now().isoformat(timespec='seconds')}",
        f"source_videos: {payload.get('source_videos', [])}",
        "",
        "[variants]",
    ]
    for item in payload.get("variants", []):
        lines.append(
            f"{item.get('variant_id')} | deleted_frames={len(item.get('deleted_frames', []))} | "
            f"cover={item.get('cover_similarity_status')}@{item.get('cover_timestamp')} | "
            f"{item.get('output_path')}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def log_event(path: Path, message: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().isoformat(timespec="seconds")
    with path.open("a", encoding="utf-8") as handle:
        handle.write(f"[{timestamp}] {message}\n")


def write_error(task_root: Path, stage: str, args: argparse.Namespace, variants: list[VariantPlan], exc: BaseException) -> None:
    payload = {
        "status": "failed",
        "stage": stage,
        "error_type": type(exc).__name__,
        "error_message": str(exc),
        "interrupted": isinstance(exc, KeyboardInterrupt),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "config": vars(args),
        "completed_count": len(variants),
        "generated_outputs": [str(item.output_path) for item in variants],
        "variants": [item.to_dict() for item in variants],
        "traceback": traceback.format_exc(),
    }
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "error.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log_event(task_root / "run.log", f"failed stage={stage} type={type(exc).__name__} message={exc} completed={len(variants)}")


if __name__ == "__main__":
    raise SystemExit(main())
