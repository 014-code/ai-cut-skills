from __future__ import annotations

import asyncio
from contextlib import AsyncExitStack
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Awaitable, Callable, Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook

from .usergrowth_browser import HOME_URL, UserGrowthBrowserClient
from .usergrowth_models import UserGrowthCancelled


ProgressCallback = Callable[[str], None]
CheckpointCallback = Callable[[dict], None]
ChunkSuccessCallback = Callable[
    ["TomatoMusicTagBatch", "TomatoMusicChunkResult"],
    Awaitable[int | None] | int | None,
]
CID_PATTERN = re.compile(r"^[0-9a-fA-F]{32}$")
MAX_SEARCH_CHUNK_SIZE = 50


@dataclass
class TomatoMusicTagBatch:
    bid: str
    tag: str
    cids: list[str]
    song_names: list[str] = field(default_factory=list)
    tracks: list[dict[str, str]] = field(default_factory=list)


@dataclass
class TomatoMusicChunkResult:
    bid: str
    tag: str
    chunk_index: int
    requested_cids: list[str]
    matched_count: int = 0
    task_id: str = ""
    total: int = 0
    success: int = 0
    failed: int = 0
    status: str = "pending"
    message: str = ""
    status_updated: bool = False
    status_updated_rows: int = 0
    status_update_message: str = ""


def normalise_bid(value: object) -> str:
    text = str(value or "").strip()
    if text.lower().startswith("bid_"):
        text = text[4:]
    return text


def tag_for_bid(value: object) -> str:
    bid = normalise_bid(value)
    return f"bid_{bid}" if bid else ""


def validate_tomato_music_tag_batches(
        batches: Iterable[TomatoMusicTagBatch],
) -> list[TomatoMusicTagBatch]:
    """限制番茄音乐工具只能为每个 BID 追加对应的 bid_<BID> 标签。"""
    validated: list[TomatoMusicTagBatch] = []
    for index, batch in enumerate(batches, start=1):
        bid = normalise_bid(batch.bid)
        expected_tag = tag_for_bid(bid)
        supplied_tag = str(batch.tag or "").strip()
        if not bid:
            raise RuntimeError(f"番茄音乐打标第 {index} 批缺少 BID，无法生成 bid_<BID> 标签")
        if supplied_tag and supplied_tag != expected_tag:
            raise RuntimeError(
                f"番茄音乐打标第 {index} 批标签不合法：BID {bid} 只能追加 {expected_tag}，"
                "红果短剧自定义标签必须通过正式上传状态机处理。"
            )
        validated.append(TomatoMusicTagBatch(
            bid=bid,
            tag=expected_tag,
            cids=list(batch.cids),
            song_names=list(batch.song_names),
            tracks=list(batch.tracks),
        ))
    return validated


def normalise_cids(values: Iterable[object]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        for token in re.split(r"[\s,，;；]+", str(value or "").strip()):
            cid = token.strip().lower()
            if not CID_PATTERN.fullmatch(cid) or cid in seen:
                continue
            seen.add(cid)
            result.append(cid)
    return result


def load_tomato_music_batches(path: Path) -> list[TomatoMusicTagBatch]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json_batches(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel_batches(path)
    raise RuntimeError(f"不支持的番茄音乐打标输入文件：{path.suffix}")


def _load_json_batches(path: Path) -> list[TomatoMusicTagBatch]:
    with path.open("r", encoding="utf-8") as fp:
        payload = json.load(fp)
    rows = payload.get("batches") if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise RuntimeError("JSON 必须是 batches 数组或包含 batches 的对象。")
    batches: list[TomatoMusicTagBatch] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        bid = normalise_bid(row.get("bid") or row.get("bookid"))
        cids = normalise_cids(row.get("cids") or [row.get("cid")])
        if not bid or not cids:
            continue
        song_names = [
            str(item).strip()
            for item in (row.get("songNames") or row.get("song_names") or [])
            if str(item).strip()
        ]
        tracks = [
            {"song": str(item.get("song") or "").strip(), "artist": str(item.get("artist") or "").strip()}
            for item in (row.get("tracks") or [])
            if isinstance(item, dict) and str(item.get("song") or "").strip() and str(item.get("artist") or "").strip()
        ]
        batches.append(
            TomatoMusicTagBatch(
                bid=bid,
                tag=str(row.get("tag") or tag_for_bid(bid)).strip(),
                cids=cids,
                song_names=song_names,
                tracks=tracks,
            )
        )
    if not batches:
        raise RuntimeError("输入 JSON 中没有有效的 BID/CID 批次。")
    return batches


def _compact_header(value: object) -> str:
    return re.sub(r"[\s_\-（）()【】\[\]]+", "", str(value or "").strip().lower())


def _load_excel_batches(path: Path) -> list[TomatoMusicTagBatch]:
    # Some exported Tomato Music workbooks expose an incomplete dimension in
    # their XML.  openpyxl's read-only iterator then yields only column A even
    # though the remaining cells are present.  Normal mode reads the actual
    # cell records and still keeps this path read-only from the user's point of
    # view (we never save the workbook).
    workbook = load_workbook(path, read_only=False, data_only=True)
    grouped: dict[str, dict[str, object]] = {}
    bid_headers = {"bid", "bookid", "书籍id", "小说id"}
    cid_headers = {"cid", "素材cid", "creativeid", "素材id"}
    song_headers = {"歌名", "歌曲名", "songname", "song"}
    status_headers = {"打标状态", "标签状态", "tagstatus", "status"}
    try:
        for sheet in workbook.worksheets:
            header_row = None
            header_map: dict[str, int] = {}
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > 30:
                    break
                compacted = [_compact_header(value) for value in row]
                bid_index = next((i for i, value in enumerate(compacted) if value in bid_headers), None)
                cid_index = next((i for i, value in enumerate(compacted) if value in cid_headers), None)
                if bid_index is None or cid_index is None:
                    continue
                header_row = row_index
                header_map = {"bid": bid_index, "cid": cid_index}
                song_index = next((i for i, value in enumerate(compacted) if value in song_headers), None)
                if song_index is not None:
                    header_map["song"] = song_index
                status_index = next((i for i, value in enumerate(compacted) if value in status_headers), None)
                if status_index is not None:
                    header_map["status"] = status_index
                break
            if header_row is None:
                continue

            last_bid = ""
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                bid_value = row[header_map["bid"]] if header_map["bid"] < len(row) else ""
                bid = normalise_bid(bid_value) or last_bid
                cid_value = row[header_map["cid"]] if header_map["cid"] < len(row) else ""
                cids = normalise_cids([cid_value])
                if not bid or not cids:
                    continue
                last_bid = bid
                status_index = header_map.get("status")
                if status_index is None:
                    inferred_index = header_map["cid"] + 1
                    inferred_status = str(row[inferred_index] or "").strip() if inferred_index < len(row) else ""
                    if inferred_status in {"已打标", "未打标"}:
                        status_index = inferred_index
                status = str(row[status_index] or "").strip() if status_index is not None and status_index < len(row) else ""
                if status == "已打标":
                    continue
                if status not in {"", "未打标"}:
                    continue
                item = grouped.setdefault(bid, {"cids": [], "songs": []})
                for cid in cids:
                    if cid not in item["cids"]:
                        item["cids"].append(cid)
                song_index = header_map.get("song")
                if song_index is not None and song_index < len(row):
                    song = str(row[song_index] or "").strip()
                    if song and song not in item["songs"]:
                        item["songs"].append(song)
    finally:
        workbook.close()

    batches = [
        TomatoMusicTagBatch(
            bid=bid,
            tag=tag_for_bid(bid),
            cids=list(values["cids"]),
            song_names=list(values["songs"]),
            tracks=[],
        )
        for bid, values in grouped.items()
        if values["cids"]
    ]
    if not batches:
        raise RuntimeError("Excel 的所有 sheet 中都没有同时找到有效的 BID 和 CID 列。")
    return batches


def split_cids(cids: list[str], chunk_size: int) -> list[list[str]]:
    size = max(1, min(int(chunk_size or MAX_SEARCH_CHUNK_SIZE), MAX_SEARCH_CHUNK_SIZE))
    return [cids[index:index + size] for index in range(0, len(cids), size)]


def serialise_results(results: list[TomatoMusicChunkResult]) -> list[dict]:
    return [asdict(item) for item in results]


class TomatoMusicTaggingClient(UserGrowthBrowserClient):
    """复用 UserGrowth 登录、全选和操作任务验收能力，为番茄音乐 CID 批量追加 BID 标签。"""

    async def run_tagging(
            self,
            batches: list[TomatoMusicTagBatch],
            *,
            customer_id: str = "",
            material_url: str = "",
            chunk_size: int = MAX_SEARCH_CHUNK_SIZE,
            progress: ProgressCallback | None = None,
            checkpoint: CheckpointCallback | None = None,
            on_chunk_success: ChunkSuccessCallback | None = None,
            playwright_instance=None,
    ) -> list[TomatoMusicChunkResult]:
        batches = validate_tomato_music_tag_batches(batches)
        try:
            from playwright.async_api import async_playwright
        except ImportError as exc:
            raise RuntimeError("需要先安装 playwright，并执行 playwright install chromium") from exc

        results: list[TomatoMusicChunkResult] = []
        async with AsyncExitStack() as playwright_stack:
            playwright = playwright_instance
            if playwright is None:
                playwright = await playwright_stack.enter_async_context(async_playwright())
            browser = await self._launch_browser(playwright)
            session = {"browser": browser}
            self._prepare_storage_state()
            context = await browser.new_context(**self._context_options())
            page = await context.new_page()
            self._wrap_page_speed(page)
            page.set_default_timeout(self.timeout_ms)
            page.set_default_navigation_timeout(self.timeout_ms)
            try:
                while True:
                    try:
                        await self._login(page, progress)
                        await self._persist_session_state(context, progress)
                        await self._enable_post_login_resource_blocking(context, progress)
                        page = await self._open_tomato_material_page(
                            page,
                            customer_id=customer_id,
                            material_url=material_url,
                            progress=progress,
                        )
                        break
                    except UserGrowthCancelled:
                        raise
                    except Exception as exc:
                        loading_stalled = await self._page_is_blank_or_loading(page)
                        if not self._is_recoverable_session_exception(exc) and not loading_stalled:
                            raise
                        page = await self._wait_for_network_recovery(
                            page,
                            context,
                            progress,
                            "番茄音乐登录和素材页准备",
                            playwright=playwright,
                            session=session,
                        )
                        context = page.context
                base_url = self._material_search_base_url(page.url)
                for batch_index, batch in enumerate(batches, start=1):
                    batch_result_start = len(results)
                    chunks = split_cids(batch.cids, chunk_size)
                    self._emit(
                        progress,
                        f"番茄音乐 BID {batch.bid}：{len(batch.cids)} 个 CID，拆成 {len(chunks)} 组搜索",
                    )
                    for chunk_index, cids in enumerate(chunks, start=1):
                        self._raise_if_cancelled()
                        result = TomatoMusicChunkResult(
                            bid=batch.bid,
                            tag=batch.tag or tag_for_bid(batch.bid),
                            chunk_index=chunk_index,
                            requested_cids=list(cids),
                        )
                        results.append(result)
                        while True:
                            self._raise_if_cancelled()
                            try:
                                page = await self._tag_one_cid_chunk(
                                    page,
                                    base_url=base_url,
                                    cids=cids,
                                    tag=result.tag,
                                    progress=progress,
                                    result=result,
                                    require_exact_match=bool(on_chunk_success),
                                )
                                break
                            except UserGrowthCancelled:
                                raise
                            except Exception as exc:
                                # Network/page failures are session failures, not
                                # a completed chunk. Recover the browser and retry
                                # this same chunk without consuming its business
                                # retry budget; a user close still raises above.
                                loading_stalled = await self._page_is_blank_or_loading(page)
                                if self._is_recoverable_session_exception(exc) or loading_stalled:
                                    page = await self._wait_for_network_recovery(
                                        page,
                                        context,
                                        progress,
                                        f"番茄音乐 BID {batch.bid} 第 {chunk_index} 组",
                                        playwright=playwright,
                                        session=session,
                                    )
                                    context = page.context
                                    page = await self._open_tomato_material_page(
                                        page,
                                        customer_id=customer_id,
                                        material_url=material_url,
                                        progress=progress,
                                    )
                                    base_url = self._material_search_base_url(page.url)
                                    self._emit(
                                        progress,
                                        f"番茄音乐 BID {batch.bid} 第 {chunk_index} 组网络恢复，"
                                        "继续当前分组",
                                    )
                                    continue
                                result.status = "failed"
                                result.message = str(exc) or "本组打标失败"
                                self._emit(
                                    progress,
                                    f"番茄音乐 BID {batch.bid} 第 {chunk_index} 组失败："
                                    f"{result.message}；已记录失败，继续后续分组",
                                )
                                if checkpoint:
                                    checkpoint({"results": serialise_results(results)})
                                break
                        if on_chunk_success and result.status == "success":
                            try:
                                callback_result = on_chunk_success(batch, result)
                                if asyncio.iscoroutine(callback_result):
                                    callback_result = await callback_result
                                result.status_updated = True
                                result.status_updated_rows = int(callback_result or 0)
                            except Exception as exc:
                                result.status = "failed"
                                result.status_update_message = str(exc)
                                result.message = f"墨攻标签任务已成功，但打标状态回写失败：{exc}"
                                if checkpoint:
                                    checkpoint({"results": serialise_results(results)})
                                self._emit(
                                    progress,
                                    f"番茄音乐 BID {batch.bid} 第 {chunk_index} 组状态回写失败："
                                    f"{result.message}；已记录失败，继续后续分组",
                                )
                        if checkpoint:
                            checkpoint({"results": serialise_results(results)})
                    batch_results = results[batch_result_start:]
                    failed_count = sum(1 for item in batch_results if item.status == "failed")
                    if failed_count:
                        self._emit(
                            progress,
                            f"番茄音乐 BID {batch.bid} 已遍历全部分组，"
                            f"成功 {len(batch_results) - failed_count} 组，失败 {failed_count} 组；"
                            "继续后续 BID 批次",
                        )
                    else:
                        self._emit(progress, f"番茄音乐 BID {batch.bid} 已完成全部可检索素材打标")
                return results
            finally:
                try:
                    await self._persist_session_state(context)
                except Exception:
                    pass
                try:
                    # 标记为流程主动收尾，避免异常退出被误记为用户关闭页面。
                    await self._close_browser_intentionally(session.get("browser", browser))
                except Exception:
                    pass

    async def _open_tomato_material_page(
            self,
            page,
            *,
            customer_id: str,
            material_url: str,
            progress: ProgressCallback | None,
    ):
        if material_url:
            await self._safe_goto(page, material_url)
            await page.wait_for_timeout(2500)
            await self._wait_tomato_material_page_ready(page)
            return page

        await self._safe_goto(page, HOME_URL)
        await page.wait_for_timeout(2500)
        if customer_id:
            page = await self._select_customer(page, customer_id, progress)

        # 客户入口会保留上次所在的墨攻页面。若已经落在带 selectorId 的
        # 素材管理页，直接复用当前页，避免再点工单管理造成错误的导航失败。
        if await self._is_tomato_material_page_ready(page):
            self._emit(progress, "客户进入后已在素材管理，直接复用当前页面")
            return page

        # 复用汽水/红果已验证的墨攻菜单导航；客户进入后必须保留当前页面，
        # 不能重新访问无 selectorId 的首页，否则会丢失客户 3681575 上下文。
        try:
            await self._open_work_order_management(page, progress, navigate_home=False)
        except RuntimeError:
            # 工单菜单的点击可能未切页，但客户入口已异步跳到素材管理。
            # 这是番茄流程的目标页，不应触发全流程清理并关闭浏览器。
            if await self._is_tomato_material_page_ready(page):
                self._emit(progress, "工单导航未切页，但当前已是素材管理，继续当前页面")
                return page
            raise

        if await self._is_tomato_material_page_ready(page):
            self._emit(progress, "已在素材管理，跳过重复菜单点击")
            return page
        page = await self._open_material_management_page(page, progress)
        await self._wait_tomato_material_page_ready(page)
        return page

    async def _is_tomato_material_page_ready(self, page) -> bool:
        body = await self._body_text(page, timeout_ms=2500)
        return "素材管理" in body and "全部素材" in body and "全局搜索" in body

    async def _wait_tomato_material_page_ready(self, page) -> None:
        async def ready() -> bool:
            return await self._is_tomato_material_page_ready(page)

        if not await self._wait_for_result(ready, timeout_ms=60000, interval_ms=800):
            await self._snapshot_error(page, "tomato_music_material_page_not_ready")
            raise RuntimeError("未进入墨攻素材管理页")

    @staticmethod
    def _material_search_base_url(url: str) -> str:
        parts = urlsplit(str(url or ""))
        params = [(key, value) for key, value in parse_qsl(parts.query, keep_blank_values=True) if key != "q"]
        query = urlencode(params)
        return urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment))

    @staticmethod
    def _material_search_url(base_url: str, cids: list[str]) -> str:
        parts = urlsplit(base_url)
        params = parse_qsl(parts.query, keep_blank_values=True)
        params.append(("q", " ".join(cids)))
        return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(params), parts.fragment))

    async def _tag_one_cid_chunk(
            self,
            page,
            *,
            base_url: str,
            cids: list[str],
            tag: str,
            progress: ProgressCallback | None,
            result: TomatoMusicChunkResult,
        require_exact_match: bool = False,
    ):
        matched_count = await self._search_redfruit_materials_by_cids(
            page,
            cids,
            progress,
            wait_on_empty=True,
            clear_query_scope=True,
        )
        result.matched_count = matched_count
        if matched_count > len(cids):
            raise RuntimeError(
                f"搜索结果 {matched_count} 条超过本组 CID 数 {len(cids)}，为避免误选已停止"
            )
        if require_exact_match and matched_count != len(cids):
            raise RuntimeError(
                f"搜索结果 {matched_count} 条少于本组 CID 数 {len(cids)}，"
                "无法安全回写逐行打标状态，已停止提交"
            )

        if await self._selected_count(page) > 0:
            await self._clear_redfruit_material_selection(page, progress)
        if not await self._click_first_material_card(page, []):
            await self._snapshot_error(page, "tomato_music_first_card_not_selected")
            raise RuntimeError("未选中番茄音乐搜索结果中的第一条素材")
        await self._wait_redfruit_material_selection_bar_visible(page)
        await self._select_redfruit_all_materials(page, matched_count)
        selected = await self._selected_count(page)
        if selected != matched_count:
            raise RuntimeError(f"实际选中 {selected} 条，预期 {matched_count} 条，已停止提交")

        await self._run_material_edit_action(page, "修改自定义标签")
        await self._fill_tomato_custom_tag_dialog(page, tag)
        before_pages = list(page.context.pages)
        before_url = page.url
        await self._submit_tomato_custom_tag_dialog(page)
        task_page = await self._wait_redfruit_arlp_task_page(page, before_pages, before_url, progress)
        task_id = await self._read_current_task_id(task_page, progress)
        result.task_id = task_id
        task_result = await self._wait_redfruit_arlp_task_result(task_page, progress, "修改自定义标签")
        result.total = int(task_result.get("total") or 0)
        result.success = int(task_result.get("success") or 0)
        result.failed = int(task_result.get("failed") or 0)
        if not self._redfruit_operation_all_expected_success(task_result, matched_count):
            raise RuntimeError(
                f"标签任务 {task_id} 未全部成功：成功 {result.success}/{result.total}，失败 {result.failed}，"
                f"目标 {matched_count} 条"
            )
        result.status = "success"
        result.message = f"任务 {task_id} 全部成功"
        self._emit(
            progress,
            f"番茄音乐 {result.bid} 第 {result.chunk_index} 组：任务 {task_id}，"
            f"成功 {result.success}/{result.total}，失败 {result.failed}",
        )
        await self._snapshot(
            task_page,
            f"tomato_music_tag_success_{result.bid}_{result.chunk_index}",
            screenshot=True,
        )
        await self._close_redfruit_result_dialog(page)
        if page.is_closed():
            page = task_page
        return page

    async def _fill_tomato_custom_tag_dialog(self, page, tag: str) -> None:
        dialog = await self._wait_tomato_custom_tag_dialog(page)
        body = await self._locator_text(dialog, timeout_ms=3000)
        if tag in body:
            return
        inputs = dialog.locator(".arco-input-tag input, input")
        input_box = None
        for index in range(await inputs.count() - 1, -1, -1):
            candidate = inputs.nth(index)
            if await candidate.is_visible():
                input_box = candidate
                break
        if input_box is None:
            raise RuntimeError("修改自定义标签弹窗中未找到标签输入框")
        await input_box.fill(tag)
        await input_box.press("Enter")
        await page.wait_for_timeout(400)
        if tag not in await self._locator_text(dialog, timeout_ms=3000):
            raise RuntimeError(f"自定义标签未成功加入弹窗：{tag}")

    async def _wait_tomato_custom_tag_dialog(self, page):
        async def find_dialog():
            candidates = page.locator(
                ".arco-modal:has-text('修改自定义标签'), "
                "[role='dialog']:has-text('修改自定义标签')"
            )
            for index in range(await candidates.count()):
                candidate = candidates.nth(index)
                if await candidate.is_visible():
                    return candidate
            return None

        dialog = await self._wait_for_result(find_dialog, timeout_ms=15000, interval_ms=300)
        if not dialog:
            raise RuntimeError("修改自定义标签弹窗未出现")
        return dialog

    async def _submit_tomato_custom_tag_dialog(self, page) -> None:
        dialog = await self._wait_tomato_custom_tag_dialog(page)
        try:
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(250)
        except Exception:
            pass
        for text in ("确定", "保存", "确认"):
            button = dialog.get_by_role("button", name=text, exact=True).first
            try:
                if await button.count() and await button.is_visible() and await button.is_enabled():
                    await button.click()
                    await page.wait_for_timeout(1800)
                    return
            except Exception:
                continue
        raise RuntimeError("修改自定义标签弹窗中未找到可点击的确定按钮")


def build_dry_run_payload(
        batches: list[TomatoMusicTagBatch],
        *,
        input_path: Path | str,
        customer_id: str,
        chunk_size: int,
) -> dict:
    return {
        "workflow": "tomato_music_bid_tagging",
        "dry_run": True,
        "input": str(input_path),
        "customer_id": customer_id,
        "chunk_size": max(1, min(chunk_size, MAX_SEARCH_CHUNK_SIZE)),
        "summary": {
            "batches": len(batches),
            "cids": sum(len(batch.cids) for batch in batches),
            "chunks": sum(len(split_cids(batch.cids, chunk_size)) for batch in batches),
        },
        "batches": [asdict(batch) for batch in batches],
    }
